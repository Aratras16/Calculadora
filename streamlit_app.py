
import streamlit as st
import pandas as pd
import io
import math
from datetime import date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google import genai
from google.genai import types as genai_types
import pypdf
from supabase import create_client
import json
from streamlit_cookies_controller import CookieController

_cookie = CookieController()

ADMIN_EMAIL = [st.secrets["admin"]["email"], st.secrets["admin"]["email2"]]


# =========================
# Base de Datos (Supabase)
# =========================

@st.cache_resource
def _get_supabase():
    return create_client(
        st.secrets["supabase"]["url"],
        st.secrets["supabase"]["key"]
    )

def _generar_folio():
    result = _get_supabase().rpc("generar_folio").execute()
    return result.data

_LIMITE_PDF = 50 * 1024 * 1024  # 50 MB

class _PDFProxy:
    """Emula UploadedFile de Streamlit para PDFs recuperados de Storage."""
    def __init__(self, data: bytes, name: str):
        self._data = data
        self.name = name
    def getvalue(self) -> bytes:
        return self._data

def _comprimir_pdf(pdf_bytes: bytes) -> bytes:
    """
    Comprime el PDF en dos pasos:
      Paso 1 (rápido): optimiza streams, imágenes y fuentes con fitz.
      Paso 2 (si sigue >50 MB): re-renderiza cada página como JPEG 120 DPI.
    """
    import fitz

    # Paso 1 — optimización lossless
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        buf = io.BytesIO()
        doc.save(buf, garbage=4, deflate=True, deflate_images=True, deflate_fonts=True, clean=True)
        doc.close()
        resultado = buf.getvalue()
    except Exception:
        resultado = pdf_bytes

    if len(resultado) <= _LIMITE_PDF:
        return resultado

    # Paso 2 — re-renderizado JPEG por página (garantiza reducción de tamaño)
    try:
        src = fitz.open(stream=resultado, filetype="pdf")
        dst = fitz.open()
        for page in src:
            mat = fitz.Matrix(120 / 72, 120 / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_bytes = pix.tobytes("jpg")
            nueva = dst.new_page(width=page.rect.width, height=page.rect.height)
            nueva.insert_image(nueva.rect, stream=img_bytes)
        buf2 = io.BytesIO()
        dst.save(buf2, garbage=4, deflate=True)
        src.close()
        dst.close()
        return buf2.getvalue()
    except Exception:
        return resultado

def _subir_pdf_storage(pdf_bytes: bytes, nombre: str) -> str:
    import unicodedata, re
    normalizado = unicodedata.normalize("NFD", nombre)
    sin_tildes = "".join(c for c in normalizado if unicodedata.category(c) != "Mn")
    path = re.sub(r"[^A-Za-z0-9_.\-]", "_", sin_tildes)
    _get_supabase().storage.from_("cotizaciones").upload(
        path, pdf_bytes, {"content-type": "application/pdf", "upsert": "true"}
    )
    return path

def _descargar_pdf_storage(path: str) -> bytes:
    return _get_supabase().storage.from_("cotizaciones").download(path)

def _url_pdf(path: str) -> str:
    result = _get_supabase().storage.from_("cotizaciones").create_signed_url(path, 7200)
    return result.get("signedURL", "") if isinstance(result, dict) else ""

def _guardar_en_db(folio, folio_origen=None):
    df = st.session_state.items_df
    total_mon = sum(m["Total c/Fee"] for m in st.session_state.monederos_list)
    if not df.empty:
        totales = df[[f"Subtotal {m}" for m in MARGINS]].sum()
        t21 = float(totales[f"Subtotal {MARGINS[0]}"]) + total_mon
        t60 = float(totales[f"Subtotal {MARGINS[-1]}"]) + total_mon
    else:
        t21 = t60 = total_mon
    pdf_path = ""
    if st.session_state.get("uploaded_pdf") is not None:
        try:
            pdf_val = st.session_state.uploaded_pdf.getvalue()
            if len(pdf_val) > _LIMITE_PDF:
                mb = len(pdf_val) / (1024 * 1024)
                st.error(
                    f"❌ El PDF ({mb:.1f} MB) supera el límite de 50 MB y no puede subirse. "
                    "Vuelve a cargar el archivo con menor resolución de imágenes."
                )
                st.stop()
            safe_name = f"{folio}_{st.session_state.pdf_filename}"
            pdf_path = _subir_pdf_storage(pdf_val, safe_name)
        except Exception as e:
            st.error(f"❌ Error al subir PDF — {type(e).__name__}: {e}")
            st.stop()
    usuario_email = st.session_state.get("user").email if st.session_state.get("user") else ""
    _get_supabase().table("cotizaciones").upsert({
        "folio": folio,
        "fecha": date.today().isoformat(),
        "hubspot_link": st.session_state.hubspot_link,
        "nombre_pdf": st.session_state.pdf_filename or "",
        "pdf_path": pdf_path,
        "modalidad": st.session_state.modalidad_global,
        "tipo_cobro": st.session_state.tarifa_global,
        "recursos_json": df.to_json(orient="records"),
        "monederos_json": json.dumps(st.session_state.monederos_list),
        "total_21": t21,
        "total_60": t60,
        "folio_origen": folio_origen,
        "usuario_email": usuario_email,
        "created_at": date.today().isoformat(),
    }).execute()

def _cargar_de_db(folio):
    result = _get_supabase().table("cotizaciones").select("*").eq("folio", folio.strip().upper()).execute()
    return result.data[0] if result.data else None

def _obtener_historial(modalidad=None, fecha_desde=None):
    query = _get_supabase().table("cotizaciones").select(
        "folio, fecha, hubspot_link, nombre_pdf, pdf_path, modalidad, tipo_cobro, total_21, total_60, folio_origen, usuario_email"
    ).order("created_at", desc=True)
    if modalidad and modalidad != "Todas":
        query = query.eq("modalidad", modalidad)
    if fecha_desde:
        query = query.gte("fecha", fecha_desde.isoformat())
    return query.execute().data

def _obtener_historial_admin(modalidad=None, fecha_desde=None, fecha_hasta=None):
    query = (_get_supabase().table("cotizaciones")
             .select("*")
             .order("created_at", desc=True))
    if modalidad:
        query = query.eq("modalidad", modalidad)
    if fecha_desde:
        query = query.gte("fecha", fecha_desde.isoformat())
    if fecha_hasta:
        query = query.lte("fecha", fecha_hasta.isoformat())
    return query.execute().data

# Definido aquí para que esté disponible tanto en admin como en la calculadora
MARGINS = ["21%", "22%", "23%", "25%", "30%", "40%", "50%", "60%"]

def generar_excel(datos, df, monederos_list=None, modalidad=None, tarifa=None):
    output = io.BytesIO()
    if monederos_list is None:
        monederos_list = []

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0E2B5C", end_color="0E2B5C", fill_type="solid")
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )
    accent_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    monedero_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    totales_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Cotización", index=False, startrow=2)

    output.seek(0)
    wb = openpyxl.load_workbook(output)
    ws = wb["Cotización"]

    _tarifa = tarifa if tarifa is not None else st.session_state.tarifa_global
    _modalidad = modalidad if modalidad is not None else st.session_state.modalidad_global
    label_tiempo_excel = "Meses" if _tarifa == "Mensual" else "Horas"
    info_texto = f"📋 DETALLE DE COTIZACIÓN | Modalidad: {_modalidad} | Cobro: {_tarifa.upper()}"

    ws.cell(row=1, column=1, value=info_texto)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['A'].width = 30

    for cell in ws["3:3"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text
        cell.border = thin_border

    for row in ws.iter_rows(min_row=4, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if cell.column >= 4:
                cell.number_format = '"$"#,##0.00'

    total_monederos_excel = 0
    if monederos_list:
        row_mon_titulo = ws.max_row + 2
        t_cell = ws.cell(row=row_mon_titulo, column=1, value="👛 MONEDEROS")
        t_cell.font = Font(bold=True, color="0E2B5C", size=11)
        t_cell.fill = accent_fill
        ws.merge_cells(start_row=row_mon_titulo, start_column=1, end_row=row_mon_titulo, end_column=6)

        mon_headers = ["Tipo", "Monto Base", "# de Monederos", "Total"]
        row_mon_header = row_mon_titulo + 1
        for ci, h in enumerate(mon_headers, start=1):
            c = ws.cell(row=row_mon_header, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            c.alignment = center_aligned_text
            c.border = thin_border

        for ri, mon in enumerate(monederos_list, start=row_mon_header + 1):
            vals = [mon["Tipo"], mon["Monto Base"], mon["Personas"], mon["Total c/Fee"]]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = thin_border
                c.fill = monedero_fill
                c.alignment = Alignment(vertical="center", horizontal="center")
                if ci in (2, 4):
                    c.number_format = '"$"#,##0.00'
            total_monederos_excel += mon["Total c/Fee"]

        row_mon_total = row_mon_header + len(monederos_list) + 1
        lbl = ws.cell(row=row_mon_total, column=3, value="TOTAL MONEDEROS")
        lbl.font = Font(bold=True, color="0E2B5C")
        lbl.alignment = Alignment(horizontal="right", vertical="center")
        lbl.fill = accent_fill
        lbl.border = thin_border
        val_mon = ws.cell(row=row_mon_total, column=4, value=total_monederos_excel)
        val_mon.number_format = '"$"#,##0.00'
        val_mon.font = Font(bold=True, size=11, color="0E2B5C")
        val_mon.fill = accent_fill
        val_mon.border = thin_border
        val_mon.alignment = center_aligned_text

    row_titulos = ws.max_row + 2
    row_valores = row_titulos + 1

    titulo_cell = ws.cell(row=row_titulos, column=1, value="RESUMEN DE TOTALES (Recursos + Monederos)")
    titulo_cell.font = Font(bold=True, color="0E2B5C", size=12)

    col_inicio_subtotales = df.columns.get_loc(f"Subtotal {MARGINS[0]}") + 1
    ws.merge_cells(start_row=row_titulos, start_column=1, end_row=row_titulos, end_column=col_inicio_subtotales - 1)

    columnas_sumar = [f"Subtotal {m}" for m in MARGINS]
    totales_sum = df[columnas_sumar].sum()

    for col_name in columnas_sumar:
        col_idx = df.columns.get_loc(col_name) + 1
        c_header = ws.cell(row=row_titulos, column=col_idx, value=f"Total {col_name.split()[-1]}")
        c_header.font = Font(bold=True, color="64748B")
        c_header.fill = totales_fill
        c_header.alignment = center_aligned_text
        c_header.border = thin_border

        valor_final = totales_sum[col_name] + total_monederos_excel
        c_val = ws.cell(row=row_valores, column=col_idx, value=valor_final)
        c_val.number_format = '"$"#,##0.00'
        c_val.font = Font(bold=True, size=12, color="1E293B")
        c_val.border = thin_border
        c_val.alignment = center_aligned_text
        c_val.fill = totales_fill

    t_min_excel = totales_sum[f"Subtotal {MARGINS[0]}"] + total_monederos_excel
    t_max_excel = totales_sum[f"Subtotal {MARGINS[-1]}"] + total_monederos_excel
    msg = f"⚠️ ADVERTENCIA: El total final (recursos + monederos) no debe ser menor (${t_min_excel:,.2f}) ({MARGINS[0]}) ni mayor (${t_max_excel:,.2f}) ({MARGINS[-1]})"
    msg_cell = ws.cell(row=row_titulos + 3, column=1, value=msg)
    msg_cell.font = Font(bold=True, color="EF4444")
    ws.merge_cells(start_row=row_titulos + 3, start_column=1, end_row=row_titulos + 3, end_column=11)

    ws.cell(row=3, column=3, value=label_tiempo_excel)

    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()

def mostrar_panel_admin():
    st.markdown("""
    <div class="admin-banner">
        <h1>🔐 Panel de Administración</h1>
        <p>Descarga las cotizaciones generadas</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns([1.2, 1, 1, 0.8])
    with col1:
        filtro_mod = st.selectbox("Modalidad", ["Todas", "DEDICADO", "STAFFING"], key="admin_filtro_mod")
    with col2:
        filtro_desde = st.date_input("Desde", value=None, key="admin_filtro_desde")
    with col3:
        filtro_hasta = st.date_input("Hasta", value=None, key="admin_filtro_hasta")
    with col4:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar", use_container_width=True, key="admin_btn_refresh"):
            st.rerun()

    historial = _obtener_historial_admin(
        modalidad=filtro_mod if filtro_mod != "Todas" else None,
        fecha_desde=filtro_desde,
        fecha_hasta=filtro_hasta,
    )

    if not historial:
        st.info("No hay cotizaciones que coincidan con los filtros.", icon="📭")
        return

    c1, c2, c3 = st.columns(3)
    c1.metric("Total cotizaciones", len(historial))
    c2.metric("DEDICADO", sum(1 for r in historial if r.get("modalidad") == "DEDICADO"))
    c3.metric("STAFFING", sum(1 for r in historial if r.get("modalidad") == "STAFFING"))

    st.markdown("---")

    rows_display = []
    for r in historial:
        pdf_url = ""
        if r.get("pdf_path"):
            try:
                pdf_url = _url_pdf(r["pdf_path"])
            except Exception:
                pdf_url = ""
        rows_display.append({
            "Folio": r["folio"],
            "Fecha": r["fecha"],
            "Usuario": r.get("usuario_email") or "-",
            "Modalidad": r.get("modalidad", ""),
            "Cobro": r.get("tipo_cobro", ""),
            "HubSpot": r.get("hubspot_link") or "",
            "PDF": r.get("nombre_pdf") or "",
            "Descargar PDF": pdf_url,
            "Total 21%": f"${r['total_21']:,.2f}" if r.get("total_21") else "-",
            "Total 60%": f"${r['total_60']:,.2f}" if r.get("total_60") else "-",
            "Basado en": r.get("folio_origen") or "-",
        })

    df_show = pd.DataFrame(rows_display)
    col_cfg = {"Descargar PDF": st.column_config.LinkColumn("Descargar PDF", display_text="📄 Ver PDF")}
    if "HubSpot" in df_show.columns:
        col_cfg["HubSpot"] = st.column_config.LinkColumn("HubSpot")
    st.dataframe(df_show, use_container_width=True, hide_index=True, column_config=col_cfg)
    st.caption(f"Total: **{len(historial)}** cotizaciones registradas.")

    st.markdown("---")

    if st.button(f"📦 Preparar ZIP completo ({len(historial)} cotizaciones)", use_container_width=True, key="btn_preparar_zip"):
        import zipfile
        cols_order = (["Rol", "Cant", "Tiempo"]
                      + [f"Precio {m}" for m in MARGINS]
                      + [f"Subtotal {m}" for m in MARGINS])
        buf_zip = io.BytesIO()
        errores = []
        with zipfile.ZipFile(buf_zip, "w", zipfile.ZIP_DEFLATED) as zf:
            progress = st.progress(0, text="Generando archivos...")
            for i, r in enumerate(historial):
                folio = r["folio"]
                carpeta = folio
                try:
                    records = json.loads(r.get("recursos_json") or "[]")
                    monederos_dl = json.loads(r.get("monederos_json") or "[]")
                    if records:
                        df_r = pd.DataFrame(records)
                        for c in [col for col in cols_order if col not in df_r.columns]:
                            df_r[c] = 0
                        df_r = df_r[[c for c in cols_order if c in df_r.columns]]
                    else:
                        df_r = pd.DataFrame(columns=cols_order)
                    xlsx_bytes = generar_excel(
                        {"Fecha de Cotizacion": r.get("fecha")},
                        df_r,
                        monederos_dl,
                        modalidad=r.get("modalidad"),
                        tarifa=r.get("tipo_cobro"),
                    )
                    zf.writestr(f"{carpeta}/{folio}.xlsx", xlsx_bytes)
                except Exception as e:
                    errores.append(f"{folio} (Excel): {e}")
                if r.get("pdf_path"):
                    try:
                        pdf_bytes = _descargar_pdf_storage(r["pdf_path"])
                        nombre_pdf = r.get("nombre_pdf") or f"{folio}.pdf"
                        zf.writestr(f"{carpeta}/{nombre_pdf}", pdf_bytes)
                    except Exception as e:
                        errores.append(f"{folio} (PDF): {e}")
                progress.progress((i + 1) / len(historial), text=f"Procesando {folio}...")
            progress.empty()
        if errores:
            st.warning("Algunos archivos no se pudieron incluir:\n" + "\n".join(errores))
        st.download_button(
            label=f"⬇️ Descargar ZIP ({len(historial)} cotizaciones)",
            data=buf_zip.getvalue(),
            file_name=f"cotizaciones_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.zip",
            mime="application/zip",
            key="admin_dl_zip",
            use_container_width=True,
            type="primary",
        )

    st.markdown("---")
    st.markdown("#### 📥 Descargar cotización individual")

    folios = [r["folio"] for r in historial]
    sel_folio = st.selectbox("Selecciona un folio", folios, key="admin_sel_folio")

    if sel_folio:
        row_data = next((r for r in historial if r["folio"] == sel_folio), None)
        if row_data:
            # 1. Datos Generales de la Cotización
            d1, d2, d3, d4 = st.columns(4)
            d1.markdown(f"**Modalidad:** {row_data.get('modalidad', '-')}")
            d2.markdown(f"**Cobro:** {row_data.get('tipo_cobro', '-')}")
            d3.markdown(f"**Folio:** `{sel_folio}`")
            d4.markdown(f"**Fecha:** {row_data.get('fecha', '-')}")

            st.markdown("---")

            # Cargar recursos y monederos del registro
            records = json.loads(row_data.get("recursos_json") or "[]")
            monederos_dl = json.loads(row_data.get("monederos_json") or "[]")
           
            df_r = pd.DataFrame()
           
            # 2. Detalle de Recursos (Tabla)
            if records:
                df_r = pd.DataFrame(records)
                label_tiempo_tabla = "Meses" if row_data.get("tipo_cobro") == "Mensual" else "Horas"
                col_cfg_recursos = {
                    "Rol": st.column_config.TextColumn("Rol/Perfil", width="medium"),
                    "Cant": st.column_config.NumberColumn("Cant.", width="small"),
                    "Tiempo": st.column_config.NumberColumn(label_tiempo_tabla, width="small"),
                }
                for m in MARGINS:
                    col_cfg_recursos[f"Precio {m}"] = None  # Ocultar
                    if m in ["21%", "25%", "60%"]:
                        col_cfg_recursos[f"Subtotal {m}"] = st.column_config.NumberColumn(f"Subtotal {m}", format="$%.2f")
                    else:
                        col_cfg_recursos[f"Subtotal {m}"] = None  # Ocultar

                cols_to_keep = ["Rol", "Cant", "Tiempo"] + [f"Precio {m}" for m in MARGINS] + [f"Subtotal {m}" for m in MARGINS]
                for c in cols_to_keep:
                    if c not in df_r.columns:
                        df_r[c] = 0.0

                st.markdown("##### Detalle de Recursos de la Cotización")
                st.dataframe(df_r[cols_to_keep], use_container_width=True, hide_index=True, column_config=col_cfg_recursos)
            else:
                st.info("Esta cotización no contiene recursos registrados.", icon="💡")

            # 4. Monederos
            if monederos_dl:
                st.markdown("##### Monederos agregados a la cotización")
                df_monederos = pd.DataFrame(monederos_dl)
                st.dataframe(df_monederos, use_container_width=True, hide_index=True)

            st.markdown("---")

            # 3. Resumen de Totales (Tarjetas de Margen)
            total_monederos_fee = sum(m["Total c/Fee"] for m in monederos_dl)
            if not df_r.empty:
                totales_r = df_r[[f"Subtotal {m}" for m in MARGINS]].sum()
            else:
                totales_r = {f"Subtotal {m}": 0.0 for m in MARGINS}

            cards_html = ""
            for m in MARGINS:
                if m in ["21%", "25%", "60%"]:
                    val_con_mon = totales_r[f"Subtotal {m}"] + total_monederos_fee
                    m_num = m.replace("%", "")
                    monedero_html = f'<div class="metric-detail">Monederos: ${total_monederos_fee:,.2f}</div>' if total_monederos_fee > 0 else ""
                   
                    cards_html += f"""
<div class="metric-container">
    <div class="metric-title">MARGEN {m}</div>
    <div class="metric-value val-{m_num}">${val_con_mon:,.2f}</div>
    <div class="metric-detail">Recursos: ${totales_r[f'Subtotal {m}']:,.2f}</div>
    {monedero_html}
</div>
"""

            html_layout = f"""
<div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 1.2rem; margin-top: 1.5rem; margin-bottom: 1.5rem;">
    {cards_html}
</div>
"""
            st.markdown(html_layout, unsafe_allow_html=True)

           

            # 5. Botones de descarga
            col_pdf, col_xls = st.columns(2)

            with col_pdf:
                if row_data.get("pdf_path"):
                    try:
                        pdf_url = _url_pdf(row_data["pdf_path"])
                        st.link_button("📄 Abrir / Descargar PDF", pdf_url, use_container_width=True)
                    except Exception:
                        st.button("📄 PDF no disponible", disabled=True, use_container_width=True, key="admin_no_pdf")
                else:
                    st.button("📄 Sin PDF adjunto", disabled=True, use_container_width=True, key="admin_no_pdf2")

            with col_xls:
                try:
                    cols_order = (["Rol", "Cant", "Tiempo"]
                                  + [f"Precio {m}" for m in MARGINS]
                                  + [f"Subtotal {m}" for m in MARGINS])
                    if not df_r.empty:
                        df_r_excel = df_r.copy()
                        for c in [col for col in cols_order if col not in df_r_excel.columns]:
                            df_r_excel[c] = 0
                        df_r_excel = df_r_excel[[c for c in cols_order if c in df_r_excel.columns]]
                    else:
                        df_r_excel = pd.DataFrame(columns=cols_order)

                    xlsx_bytes = generar_excel(
                        {"Fecha de Cotizacion": row_data.get("fecha")},
                        df_r_excel,
                        monederos_dl,
                        modalidad=row_data.get("modalidad"),
                        tarifa=row_data.get("tipo_cobro"),
                    )
                    st.download_button(
                        label=f"📊 Descargar Excel — {sel_folio}",
                        data=xlsx_bytes,
                        file_name=f"{sel_folio}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="admin_dl_excel_cot",
                        use_container_width=True,
                        type="primary",
                    )
                except Exception as e:
                    st.error(f"No se pudo generar el Excel: {e}")

# =========================
# Configuración de página
# =========================
st.set_page_config(page_title="Cotizador UX/UI", page_icon="🧮", layout="wide", initial_sidebar_state="expanded")

_get_supabase()

# =========================
# Estilos CSS Avanzados (Tema Claro)
# =========================
def inyectar_css():
    st.markdown("""
        <style>
        /* Importar fuente moderna y corporativa */
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

        /* Variables globales (Light Theme base) */
        :root {
            --primary-color: #0E2B5C;      /* Azul fuerte corporativo */
            --secondary-color: #3B82F6;    /* Azul brillante */
            --accent-color: #10B981;       /* Verde acento */
            --bg-color: #F8FAFC;           /* Fondo general más suave que el blanco puro */
            --card-bg: #FFFFFF;            /* Fondo de tarjetas */
            --text-main: #1E293B;          /* Texto oscuro para legibilidad */
            --text-muted: #64748B;         /* Texto secundario */
            --border-color: #E2E8F0;       /* Bordes muy sutiles */
        }

        /* Estilo base de Streamlit */
        .stApp {
            background-color: var(--bg-color);
            font-family: 'Inter', sans-serif !important;
            color: var(--text-main);
        }

        h1, h2, h3, h4, h5, h6, .stMarkdown, .stText, p, label, li {
            font-family: 'Inter', sans-serif !important;
        }

        h4, h5, h6, .stMarkdown, .stText, p, label, li {
            color: var(--text-main) !important;
        }

        h1, h2, h3 {
            color: var(--primary-color) !important;
        }

        /* Botones */
        button[kind="primary"] {
            background: linear-gradient(135deg, var(--secondary-color), var(--primary-color)) !important;
            color: white !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: 600 !important;
            padding: 0.6rem 1.2rem !important;
            transition: all 0.3s ease !important;
            box-shadow: 0 4px 6px -1px rgba(59, 130, 246, 0.2), 0 2px 4px -1px rgba(59, 130, 246, 0.1) !important;
        }
        button[kind="primary"] p,
        button[kind="primary"] span,
        button[kind="primary"] div {
            color: white !important;
            -webkit-text-fill-color: white !important;
        }

        button[kind="primary"]:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 10px 15px -3px rgba(59, 130, 246, 0.3), 0 4px 6px -2px rgba(59, 130, 246, 0.15) !important;
            opacity: 0.95 !important;
        }

        button[kind="secondary"] {
            background: rgba(255, 255, 255, 0.5) !important;
            color: var(--text-main) !important;
            border: 1px solid var(--border-color) !important;
            border-radius: 8px !important;
            font-weight: 500 !important;
            transition: all 0.3s ease !important;
        }
        button[kind="secondary"] p,
        button[kind="secondary"] span,
        button[kind="secondary"] div {
            color: var(--text-main) !important;
            -webkit-text-fill-color: var(--text-main) !important;
        }

        button[kind="secondary"]:hover {
            border-color: var(--secondary-color) !important;
            color: var(--secondary-color) !important;
            background: rgba(59, 130, 246, 0.05) !important;
            transform: translateY(-1px) !important;
        }
        button[kind="secondary"]:hover p,
        button[kind="secondary"]:hover span,
        button[kind="secondary"]:hover div {
            color: var(--secondary-color) !important;
            -webkit-text-fill-color: var(--secondary-color) !important;
        }

        /* Inputs de textos, selectbox y fechas */
        .stTextInput input, .stTextArea textarea, .stDateInput input, .stSelectbox select, .stNumberInput input, div[data-baseweb="select"] > div {
            border-radius: 6px !important;
            border: 1px solid var(--border-color) !important;
            transition: border-color 0.2s, box-shadow 0.2s !important;
            background-color: var(--card-bg) !important;
            color: var(--text-main) !important;
            -webkit-text-fill-color: var(--text-main) !important;
        }

        .stTextInput input:focus, .stTextArea textarea:focus, .stDateInput input:focus, .stSelectbox select:focus, .stNumberInput input:focus {
            border-color: var(--secondary-color) !important;
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.2) !important;
            outline: none !important;
        }

        /* Expander Title */
        .streamlit-expanderHeader {
            font-weight: 600 !important;
            color: var(--primary-color) !important;
            font-size: 1.1rem !important;
            background-color: var(--card-bg) !important;
            border-radius: 8px !important;
        }

        /* Tarjeta de Métricas custom */
        .metric-container {
            background-color: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 1.8rem;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            text-align: center;
        }
       
        .metric-container:hover {
            transform: translateY(-4px);
            box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1);
            border-color: var(--secondary-color);
        }

        .metric-title {
            font-size: 0.95rem;
            color: var(--text-muted);
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.06em;
            margin-bottom: 0.6rem;
        }

        .metric-value {
            font-size: 2rem;
            font-weight: 800;
            line-height: 1.2;
        }

        .metric-detail {
            font-size: 0.8rem;
            color: #94A3B8;
            margin-top: 0.3rem;
            font-weight: 500;
        }

        /* Colores semánticos sutiles pero claros */
        .val-21 { color: #6366F1 !important; }  /* Indigo */
        .val-22 { color: #3B82F6 !important; }  /* Azul */
        .val-23 { color: #10B981 !important; }  /* Verde */
        .val-25 { color: #F59E0B !important; }  /* Naranja */
        .val-30 { color: #EF4444 !important; }  /* Rojo */
        .val-40 { color: #8B5CF6 !important; }  /* Violeta */
        .val-50 { color: #EC4899 !important; }  /* Rosa */
        .val-60 { color: #1E293B !important; }  /* Oscuro */

        /* Resaltar cabecera / Banner */
        .hero-banner {
            background: linear-gradient(120deg, var(--card-bg) 0%, #E0F2FE 100%);
            padding: 2.5rem;
            border-radius: 16px;
            margin-bottom: 2rem;
            border-left: 8px solid var(--secondary-color);
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);
        }
        .hero-banner h1 {
            color: var(--primary-color) !important;
            margin-top: 0 !important;
            font-size: 2.4rem;
            font-weight: 800 !important;
            margin-bottom: 0.5rem;
        }
        .hero-banner p {
            color: var(--text-muted);
            font-size: 1.15rem;
            font-weight: 400;
            margin-bottom: 0;
        }

        /* Banner del Panel de Admin */
        .admin-banner {
            background: linear-gradient(135deg, #0E2B5C 0%, #1e40af 100%) !important;
            padding: 2rem 2.5rem !important;
            border-radius: 12px !important;
            margin-bottom: 2rem !important;
        }
        .admin-banner h1 {
            color: #ffffff !important;
            -webkit-text-fill-color: #ffffff !important;
            margin: 0 !important;
            font-size: 2rem !important;
            font-weight: 800 !important;
        }
        .admin-banner p {
            color: #ffffff !important;
            -webkit-text-fill-color: #ffffff !important;
            margin: 0.5rem 0 0 !important;
            opacity: 0.85 !important;
            font-size: 1rem !important;
        }

        /* Sección superior de sesión */
        .session-header {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 1.5rem;
            margin-bottom: 1.5rem;
        }

        /* Tarjeta de folio */
        .folio-card {
            padding: 1.2rem 1.8rem;
            border-radius: 14px;
            box-shadow: 0 4px 14px rgba(0,0,0,0.18);
        }
        .folio-card-label {
            font-size: 0.68rem;
            color: rgba(255,255,255,0.65);
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.1em;
            margin-bottom: 0.35rem;
        }
        .folio-card-value {
            font-size: 1.6rem;
            color: white;
            font-weight: 800;
            letter-spacing: 0.06em;
            font-family: 'Courier New', monospace;
        }

        /* Notificación interna */
        .badge-interno-box {
            background: #F0FDF4;
            border: 1px solid #6EE7B7;
            border-radius: 10px;
            padding: 1rem 1.2rem;
            margin-top: 1.5rem;
        }

        /* Section divider label */
        .section-label {
            font-size: 0.7rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: var(--text-muted);
            margin-bottom: 0.5rem;
        }
        </style>
    """, unsafe_allow_html=True)

def mostrar_login():
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        st.markdown("""
        <div style="text-align:center;padding:2rem 0 1.5rem;">
            <div style="font-size:3rem;">🧮</div>
            <h2 style="color:#0E2B5C;font-weight:700;margin:0.5rem 0 0.2rem;">Cotizador UX/UI</h2>
            <p style="color:#64748B;font-size:0.9rem;margin:0;">Ingresa con tu cuenta para continuar</p>
        </div>
        <div style="height:4px;background:linear-gradient(90deg,#0E2B5C,#3B82F6);border-radius:2px;margin-bottom:2rem;"></div>
        """, unsafe_allow_html=True)

        if "modo_reset" not in st.session_state:
            st.session_state.modo_reset = False

        if not st.session_state.modo_reset:
            email = st.text_input("Correo electrónico", placeholder="Ingresa tu correo (usuario@elektra.com.mx)", key="login_email")
            password = st.text_input("Contraseña", placeholder="Ingresa tu contraseña (numero de empleado)", type="password", key="login_password")
            if st.button("Iniciar sesión", type="primary", use_container_width=True, key="btn_login"):
                if not email or not password:
                    st.error("Ingresa tu correo y contraseña.")
                else:
                    try:
                        resp = _get_supabase().auth.sign_in_with_password({"email": email, "password": password})
                        st.session_state.user = resp.user
                        if resp.session:
                            # Defer cookie writing to the next run (writing cookies and calling
                            # st.rerun() in the same run causes the cookie write to be discarded)
                            st.session_state._pending_tokens = (
                                resp.session.access_token,
                                resp.session.refresh_token,
                            )
                        st.rerun()
                    except Exception:
                        st.error("Correo o contraseña incorrectos.")
            st.markdown("<br>", unsafe_allow_html=True)
            #if st.button("¿Olvidaste tu contraseña?", use_container_width=True, key="btn_ir_reset"):
             #   st.session_state.modo_reset = True
              #  st.rerun()
       

inyectar_css()

if "user" not in st.session_state:
    _access = _cookie.get('sb_access_token')
    _refresh = _cookie.get('sb_refresh_token')
    if _access and _refresh:
        try:
            _resp = _get_supabase().auth.set_session(_access, _refresh)
            if _resp and _resp.user:
                st.session_state.user = _resp.user
        except Exception:
            _cookie.remove('sb_access_token')
            _cookie.remove('sb_refresh_token')

if "user" not in st.session_state:
    mostrar_login()
    st.stop()

if "_pending_tokens" in st.session_state:
    _at, _rt = st.session_state.pop("_pending_tokens")
    _cookie.set('sb_access_token', _at)
    _cookie.set('sb_refresh_token', _rt)

if "vista_actual" not in st.session_state:
    st.session_state.vista_actual = "calculadora"

_usuario_actual = st.session_state.user

with st.sidebar:
    st.markdown(f"### Hola {_usuario_actual.email.split('@')[0].split('.')[0].title()} 👋")
    if _usuario_actual.email in ADMIN_EMAIL:
        _nc1, _nc2 = st.columns(2)
        with _nc1:
            if st.button("🧮 Cotizador", use_container_width=True, key="nav_calc",
                         type="primary" if st.session_state.vista_actual == "calculadora" else "secondary"):
                st.session_state.vista_actual = "calculadora"
                st.rerun()
        with _nc2:
            if st.button("🔐 Admin", use_container_width=True, key="nav_admin",
                         type="primary" if st.session_state.vista_actual == "admin" else "secondary"):
                st.session_state.vista_actual = "admin"
                st.rerun()
    if st.button("Cerrar sesión", key="btn_logout_nav", use_container_width=True):
        _get_supabase().auth.sign_out()
        del st.session_state["user"]
        _cookie.remove('sb_access_token')
        _cookie.remove('sb_refresh_token')
        st.rerun()
    st.divider()

if st.session_state.vista_actual == "admin" and _usuario_actual.email in ADMIN_EMAIL:
    mostrar_panel_admin()
    st.stop()

# =========================
# Cabecera Visual (Hero)
# =========================
st.markdown("""
<div class="hero-banner">
    <h1>🧮 Cotizador de Servicios UX/UI</h1>
    <p>Cálculo estructurado con márgenes de contribución para la planeación de proyectos de diseño.</p>
</div>
""", unsafe_allow_html=True)

st.markdown("<div style='background-color:#FFF3CD;padding:1rem;border-radius:8px;border-left:5px solid #FFC107;margin-bottom:2rem;'><h4><strong>⚠️ Para cualquier recotizacion o ajuste de precios es necesario ponerse en contacto con finanzas y obtener el VoBo</strong></h4></div>", unsafe_allow_html=True)


# =========================
# Historial de Cotizaciones
# =========================
st.markdown("### 📋  Historial de Cotizaciones")
with st.expander("Ver historial de cotizaciones guardadas", expanded=False):
    col_h1, col_h2, col_h3 = st.columns([1, 1, 1], gap="medium")
    with col_h1:
        filtro_modalidad = st.selectbox("Modalidad", ["Todas", "DEDICADO", "STAFFING"], key="hist_modalidad")
    with col_h2:
        filtro_fecha = st.date_input("Desde la fecha", value=None, key="hist_fecha")
    with col_h3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar", use_container_width=True, key="btn_refresh_hist"):
            st.rerun()

    historial = _obtener_historial(
        modalidad=filtro_modalidad if filtro_modalidad != "Todas" else None,
        fecha_desde=filtro_fecha,
    )

    if historial:
        for row in historial:
            if row.get("pdf_path"):
                try:
                    row["pdf_url"] = _url_pdf(row["pdf_path"])
                except Exception:
                    row["pdf_url"] = ""
            else:
                row["pdf_url"] = ""
        df_hist = pd.DataFrame(historial)
        df_hist.rename(columns={
            "folio": "Folio", "fecha": "Fecha", "hubspot_link": "HubSpot",
            "nombre_pdf": "PDF", "pdf_url": "Descargar PDF", "modalidad": "Modalidad",
            "tipo_cobro": "Cobro", "total_21": "Total 21%", "total_60": "Total 60%",
            "folio_origen": "Basado en"
        }, inplace=True)
        if "pdf_path" in df_hist.columns:
            df_hist.drop(columns=["pdf_path"], inplace=True)
        df_hist["Total 21%"] = df_hist["Total 21%"].apply(lambda x: f"${x:,.2f}" if x else "-")
        df_hist["Total 60%"] = df_hist["Total 60%"].apply(lambda x: f"${x:,.2f}" if x else "-")
        df_hist["Basado en"] = df_hist["Basado en"].fillna("-")
        col_cfg = {"Descargar PDF": st.column_config.LinkColumn("Descargar PDF", display_text="📄 Ver PDF")}
        if "HubSpot" in df_hist.columns:
            col_cfg["HubSpot"] = st.column_config.LinkColumn("HubSpot")
        st.dataframe(df_hist, use_container_width=True, hide_index=True, column_config=col_cfg)
        st.caption(f"Total: **{len(historial)}** cotizaciones registradas.")
    else:
        st.info("No hay cotizaciones guardadas que coincidan con los filtros.", icon="📭")





# =========================
# Catálogo Estructurado
# Índices: [0]=22%, [1]=23%, [2]=25%, [3]=30%
# =========================
CATALOGO = {
    "DEDICADO": {
        "DISEÑADOR UX/UI": [126455, 127500, 128545, 130635, 135861, 146311, 156762, 167213],
        "PRODUCT DESIGNER": [131199, 132283, 133367, 135536, 140957, 151800, 162643, 173486],
        "SERVICE DESIGNER": [146500, 147711, 148921, 151343, 157397, 169504, 181612, 193719],
        "CUSTOMER SUCCESS": [165410, 166777, 168144, 170878, 177713, 191384, 205054, 218724]
    },
 "STAFFING": {
    "DISEÑADOR UX/UI": [106123, 107000, 107877, 109631, 114016, 122787, 131557, 140328],
    "PRODUCT DESIGNER": [111578, 112500, 113422, 115266, 119877, 129098, 138320, 147541],
    "SERVICE DESIGNER": [119016, 120000, 120984, 122951, 127869, 137705, 147541, 157377],
    "CUSTOMER SUCCESS": [132902, 134000, 135098, 137295, 142787, 153770, 164754, 175738]
}
}

MONEDEROS = {
    "Tiendas Neto" : {
        "Monto": [200,300,400,500],
        "Monto con fee" : [200*1.05,300*1.05,400*1.05,500*1.05]
    },
    "Externo" : {
        "Monto": [200,300,400,500],
        "Monto con fee" : [200*1.15,300*1.15,400*1.15,500*1.15]
    }

}

GEMINI_TOOLS = genai_types.Tool(
    function_declarations=[
        genai_types.FunctionDeclaration(
            name="agregar_recurso",
            description="Agrega un especialista a la cotización con el rol, cantidad de personas y tiempo indicados.",
            parameters=genai_types.Schema(
                type="OBJECT",
                properties={
                    "rol": genai_types.Schema(
                        type="STRING",
                        enum=["DISEÑADOR UX/UI", "PRODUCT DESIGNER", "SERVICE DESIGNER", "CUSTOMER SUCCESS"],
                        description="Rol del especialista"
                    ),
                    "cantidad": genai_types.Schema(
                        type="INTEGER",
                        description="Número de personas con este rol"
                    ),
                    "tiempo": genai_types.Schema(
                        type="NUMBER",
                        description="Tiempo en meses (tarifa Mensual) o en horas (tarifa Por Hora)"
                    ),
                },
                required=["rol", "cantidad", "tiempo"]
            )
        ),
        genai_types.FunctionDeclaration(
            name="limpiar_recursos",
            description="Elimina todos los recursos actuales de la cotización.",
            parameters=genai_types.Schema(type="OBJECT", properties={})
        ),
    ]
)

# =========================
# Estado inicial (Session State)
# =========================
if "items_df" not in st.session_state:
    cols = ["Rol", "Cant", "Tiempo"] + [f"Precio {m}" for m in MARGINS] + [f"Subtotal {m}" for m in MARGINS]
    st.session_state.items_df = pd.DataFrame(columns=cols)

if "datos" not in st.session_state:
    st.session_state.datos = {
        "Fecha de Cotizacion": date.today(),
    }

if "uploaded_pdf" not in st.session_state:
    st.session_state.uploaded_pdf = None

if "hubspot_link" not in st.session_state:
    st.session_state.hubspot_link = ""

if "modalidad_global" not in st.session_state:
    st.session_state.modalidad_global = "DEDICADO"

if "tarifa_global" not in st.session_state:
    st.session_state.tarifa_global = "Mensual"

if "monederos_list" not in st.session_state:
    st.session_state.monederos_list = []  # lista de dicts: {tipo, monto, monto_fee, personas}

if "chat_messages" not in st.session_state:
    st.session_state.chat_messages = []

if "pdf_text" not in st.session_state:
    st.session_state.pdf_text = ""

if "pdf_filename" not in st.session_state:
    st.session_state.pdf_filename = ""

if "attachment_key" not in st.session_state:
    st.session_state.attachment_key = 0

if "token_totals" not in st.session_state:
    st.session_state.token_totals = {"entrada": 0, "salida": 0}

if "folio_actual" not in st.session_state:
    st.session_state.folio_actual = None

if "folio_origen" not in st.session_state:
    st.session_state.folio_origen = None

if "tipo_cliente" not in st.session_state:
    st.session_state.tipo_cliente = "Externo"

if "folio_preview" not in st.session_state:
    st.session_state.folio_preview = _generar_folio()

def extraer_texto_pdf(pdf_bytes: bytes) -> str:
    try:
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        return "\n".join(page.extract_text() or "" for page in reader.pages).strip()
    except Exception:
        return ""


def recalcular(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    # Asegurar tipos numéricos
    cols_num = ["Cant", "Tiempo"] + [f"Precio {m}" for m in MARGINS]
    for col in cols_num:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Recalcular totales: Precio * Cantidad * Tiempo
    for m in MARGINS:
        df[f"Subtotal {m}"] = (df[f"Precio {m}"] * df["Cant"] * df["Tiempo"]).apply(math.ceil)
    return df


def _ejecutar_herramienta(tool_name: str, tool_input: dict) -> str:
    if tool_name == "agregar_recurso":
        rol = tool_input["rol"]
        cantidad = tool_input["cantidad"]
        tiempo = tool_input["tiempo"]
        precios = CATALOGO[st.session_state.modalidad_global][rol]
        if st.session_state.tarifa_global == "Por Hora":
            precios = [p / 160.0 for p in precios]
        factor = cantidad * tiempo
        data_nueva = {"Rol": rol, "Cant": int(cantidad), "Tiempo": float(tiempo)}
        for i, m in enumerate(MARGINS):
            data_nueva[f"Precio {m}"] = precios[i]
            data_nueva[f"Subtotal {m}"] = math.ceil(precios[i] * factor)
        nuevo = pd.DataFrame([data_nueva])
        st.session_state.items_df = pd.concat([st.session_state.items_df, nuevo], ignore_index=True)
        label = "meses" if st.session_state.tarifa_global == "Mensual" else "horas"
        return f"Agregado: {cantidad}x {rol} por {tiempo} {label}"
    elif tool_name == "limpiar_recursos":
        cols = ["Rol", "Cant", "Tiempo"] + [f"Precio {m}" for m in MARGINS] + [f"Subtotal {m}" for m in MARGINS]
        st.session_state.items_df = pd.DataFrame(columns=cols)
        return "Recursos limpiados"
    return "Herramienta no reconocida"


def call_ai_agent(user_message: str, history: list, attachments: list = None):
    try:
        api_key = st.secrets["gemini"]["api_key"]
    except Exception:
        return "⚠️ Agrega `[gemini] api_key` en los secrets de Streamlit para usar el asistente.", False

    client = genai.Client(api_key=api_key)
    label_tiempo = "meses" if st.session_state.tarifa_global == "Mensual" else "horas"

    if not st.session_state.items_df.empty:
        recursos_lines = [
            f"- {int(r['Cant'])}x {r['Rol']} por {r['Tiempo']} {label_tiempo}"
            for _, r in st.session_state.items_df.iterrows()
        ]
        recursos_ctx = "\n".join(recursos_lines)
    else:
        recursos_ctx = "(ninguno)"

    system_prompt = f"""Eres un asistente experto en cotización de servicios de diseño UX/UI para Elektra, una empresa mexicana de retail y servicios financieros. Tu trabajo es ayudar a los project managers a armar cotizaciones precisas para proyectos digitales internos.

## CONTEXTO DEL NEGOCIO

El cliente contrata especialistas de diseño bajo dos modalidades:
- **DEDICADO**: el especialista es supervisado por nosotros Tarifas más altas. Ideal para proyectos de alta prioridad, plazos cortos o que requieren enfoque total.
- **STAFFING**: el especialista es supervisado por el cliente. Tarifas más bajas. Ideal para proyectos de mediana complejidad o cuando el presupuesto es ajustado.

Modalidad activa: **{st.session_state.modalidad_global}**
Tipo de tarifa activa: **{st.session_state.tarifa_global}** (tiempo en {label_tiempo})

## ROLES DISPONIBLES Y CUÁNDO USARLOS

- **DISEÑADOR UX/UI**: Diseña pantallas, flujos y componentes visuales. Se usa en proyectos de aplicaciones móviles, portales web, rediseños de interfaz, sistemas de diseño.
- **PRODUCT DESIGNER**: Combina diseño y estrategia de producto. Se usa cuando el proyecto necesita definir funcionalidades, priorización y visión de producto además del diseño.
- **SERVICE DESIGNER**: Diseña la experiencia completa del servicio, incluyendo procesos internos, touchpoints físicos y digitales. Se usa en proyectos de transformación de servicio, customer journeys, proyectos omnicanal.
- **CUSTOMER SUCCESS**: Garantiza la adopción y satisfacción del cliente con el producto entregado. Se usa en proyectos que necesitan acompañamiento post-lanzamiento, capacitación o gestión de stakeholders.

## ESTADO ACTUAL DE LA COTIZACIÓN

Recursos agregados:
{recursos_ctx}

## REGLAS DE NEGOCIO

1. **Margen mínimo (21%)**: es el precio de piso. Nunca cotizar por debajo.
2. **Margen máximo (60%)**: es el precio techo. Nunca cotizar por encima.
3. El cliente recibe un rango entre el total al 21% y al 60%, y elige dónde negociar.
4. Proyectos típicos duran entre 1 y 12 meses. Menos de 1 mes es inusual; más de 12, considerar renovación.
5. Para proyectos grandes (más de 5 personas), es común mezclar roles complementarios.

## COMPOSICIONES DE EQUIPO TÍPICAS

- **App móvil simple**: 1-2 Diseñadores UX/UI por 2-4 meses
- **App móvil compleja / portal web**: 1 Product Designer + 1-2 Diseñadores UX/UI por 3-6 meses
- **Transformación de servicio / omnicanal**: 1 Service Designer + 1 Product Designer + 1-2 Diseñadores UX/UI por 4-8 meses
- **Rediseño con adopción**: cualquier equipo + 1 Customer Success por la duración del proyecto
- **Discovery / investigación**: 1 Service Designer o Product Designer por 1-2 meses
- **Sistema de diseño**: 1-2 Diseñadores UX/UI por 2-4 meses

## INSTRUCCIONES DE COMPORTAMIENTO

1. Cuando el usuario describa un proyecto, interpreta el tipo de trabajo, complejidad y duración implícita.
2. Si la descripción es vaga, haz UNA sola pregunta clave para aclarar lo más importante antes de cotizar (duración o tipo de proyecto, nunca ambas a la vez).
3. Usa las herramientas para agregar los recursos recomendados automáticamente.
4. Si el usuario pide "limpiar", "reiniciar" o "empezar de nuevo", usa limpiar_recursos antes de agregar los nuevos.
5. Al confirmar lo que agregaste, muestra un resumen con formato: "Rol × cantidad por N {label_tiempo}".
6. Si el usuario menciona un presupuesto, oriéntalo hacia la configuración que mejor se ajuste dentro de los márgenes de negocio.
7. Responde siempre en español, de forma concisa y profesional. Sin listas innecesarias si la respuesta puede ser una oración."""

    if st.session_state.pdf_text:
        texto_pdf = st.session_state.pdf_text[:10000]
        truncado = "\n[... documento truncado ...]" if len(st.session_state.pdf_text) > 10000 else ""
        system_prompt += f"""

## DOCUMENTO DEL PROYECTO

El usuario subió el siguiente documento de proyecto. Es tu fuente principal de contexto: extrae de aquí el alcance, entregables, fases, duración estimada y cualquier mención de perfiles o equipos para hacer recomendaciones precisas.

{texto_pdf}{truncado}"""

    gemini_history = [
        genai_types.Content(
            role="user" if msg["role"] == "user" else "model",
            parts=[genai_types.Part.from_text(text=msg["content"])]
        )
        for msg in history
    ]

    did_update = False

    try:
        chat = client.chats.create(
            model= "gemini-3.1-pro-preview",
            config=genai_types.GenerateContentConfig(
                tools=[GEMINI_TOOLS],
                system_instruction=system_prompt
            ),
            history=gemini_history
        )

        user_parts = [genai_types.Part.from_text(text=user_message)]
        if attachments:
            for att in attachments:
                user_parts.append(
                    genai_types.Part.from_bytes(data=att["data"], mime_type=att["type"])
                )

        response = chat.send_message(user_parts)

        for _ in range(5):
            try:
                parts = response.candidates[0].content.parts
            except (IndexError, AttributeError):
                break

            fn_parts = [p for p in parts if p.function_call and p.function_call.name]
            if not fn_parts:
                break

            result_parts = []
            for p in fn_parts:
                fn = p.function_call
                result = _ejecutar_herramienta(fn.name, dict(fn.args))
                result_parts.append(
                    genai_types.Part.from_function_response(
                        name=fn.name,
                        response={"result": result}
                    )
                )
                did_update = True

            response = chat.send_message(result_parts)

        try:
            final_text = response.text
        except Exception:
            final_text = "Cotización actualizada."

        try:
            usage = response.usage_metadata
            st.session_state.token_totals["entrada"] += usage.prompt_token_count or 0
            st.session_state.token_totals["salida"] += usage.candidates_token_count or 0
        except Exception:
            pass

        return final_text or "Cotización actualizada.", did_update

    except Exception as e:
        return f"Error al conectar con el asistente: {e}", False

# =========================
# Sidebar: Asistente de IA
# =========================
# Cambiar a True si se desea mostrar el asistente de IA en el sidebar
MOSTRAR_ASISTENTE = False

if MOSTRAR_ASISTENTE:
    with st.sidebar:
        st.markdown("## 🤖 Asistente de Cotización")
        st.caption("Describe tu proyecto en lenguaje natural y configuraré los recursos automáticamente.")
        st.divider()

        for msg in st.session_state.chat_messages:
            avatar = "🧑‍💼" if msg["role"] == "user" else "✨"
            with st.chat_message(msg["role"], avatar=avatar):
                st.markdown(msg["content"])
                for fname in msg.get("files", []):
                    ext = fname.split(".")[-1].lower()
                    icono = "🖼️" if ext in ("png", "jpg", "jpeg", "webp") else "📄"
                    st.caption(f"{icono} {fname}")

        archivos = st.file_uploader(
            "📎 Adjuntar imágenes o PDFs al mensaje",
            type=["pdf", "png", "jpg", "jpeg", "webp"],
            accept_multiple_files=True,
            key=f"uploader_{st.session_state.attachment_key}",
        )
        if archivos:
            st.caption(f"📎 {len(archivos)} archivo(s) listo(s) para enviar")

        if user_input := st.chat_input("Ej: 2 UX designers por 3 meses...", key="agent_chat_input"):
            attachments = [
                {"name": f.name, "data": f.getvalue(), "type": f.type}
                for f in (archivos or [])
            ]
            history = list(st.session_state.chat_messages)
            st.session_state.chat_messages.append({
                "role": "user",
                "content": user_input,
                "files": [a["name"] for a in attachments]
            })
            response_text, _ = call_ai_agent(user_input, history, attachments)
            st.session_state.chat_messages.append({"role": "assistant", "content": response_text})
            if attachments:
                st.session_state.attachment_key += 1
            st.rerun()

        if st.session_state.chat_messages:
            st.divider()
            if st.button("🗑️ Limpiar chat", use_container_width=True, key="btn_clear_chat"):
                st.session_state.chat_messages = []
                st.session_state.token_totals = {"entrada": 0, "salida": 0}
                st.rerun()

        totals = st.session_state.token_totals
        total = totals["entrada"] + totals["salida"]
        if total > 0:
            st.divider()
            st.caption("📊 **Tokens usados en esta sesión**")
            col_e, col_s = st.columns(2)
            col_e.metric("Entrada", f"{totals['entrada']:,}")
            col_s.metric("Salida", f"{totals['salida']:,}")
            st.caption(f"Total: **{total:,}** tokens")


# =========================
# 0) Encabezado de sesión: Folio · Tipo de cliente · Cargar
# =========================
folio_display = st.session_state.folio_actual or st.session_state.folio_preview
es_guardado = bool(st.session_state.folio_actual)
grad_start = "#065F46" if es_guardado else "#0E2B5C"
grad_end   = "#059669" if es_guardado else "#1d4ed8"
folio_icon = "✅" if es_guardado else "📋"
folio_label_txt = "Folio guardado" if es_guardado else "Folio a generar"

col_top1, col_top2 = st.columns([1, 1], gap="large")

with col_top1:
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,{grad_start} 0%,{grad_end} 100%);
                padding:1.2rem 1.8rem;border-radius:14px;margin-bottom:1rem;
                box-shadow:0 4px 14px rgba(0,0,0,0.18);">
        <div style="font-size:0.68rem;color:rgba(255,255,255,0.65);font-weight:700;
                    text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.35rem;">
            {folio_icon} {folio_label_txt}
        </div>
        <div style="font-size:1.6rem;color:white;font-weight:800;
                    letter-spacing:0.06em;font-family:'Courier New',monospace;">
            {folio_display}
        </div>
    </div>
    """, unsafe_allow_html=True)

    if es_guardado:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🆕 Nueva cotización", use_container_width=True, key="btn_nueva_cot"):
            cols_empty = ["Rol", "Cant", "Tiempo"] + [f"Precio {m}" for m in MARGINS] + [f"Subtotal {m}" for m in MARGINS]
            st.session_state.folio_actual = None
            st.session_state.folio_origen = None
            st.session_state.folio_preview = _generar_folio()
            st.session_state.items_df = pd.DataFrame(columns=cols_empty)
            st.session_state.monederos_list = []
            st.session_state.hubspot_link = ""
            st.session_state.uploaded_pdf = None
            st.session_state.pdf_filename = ""
            st.session_state.pdf_text = ""
            st.rerun()

with col_top2:
    st.markdown("<p class='section-label'>Ingresa folio para recotización</p>", unsafe_allow_html=True)
    col_f1, col_f2 = st.columns([3, 1], gap="small")
    with col_f1:
        folio_input = st.text_input(
            "Folio",
            placeholder="COT-MMDDYYYY-XXX",
            key="input_folio_buscar",
            label_visibility="collapsed"
        )
    with col_f2:
        if st.button("Obtener datos", use_container_width=True, key="btn_cargar_folio"):
            if folio_input.strip():
                data = _cargar_de_db(folio_input.strip())
                if data:
                    st.session_state.modalidad_global = data["modalidad"]
                    st.session_state.tarifa_global = data["tipo_cobro"]
                    st.session_state.hubspot_link = data["hubspot_link"] or ""
                    st.session_state.tipo_cliente = "Externo" if data.get("hubspot_link") else "Interno"
                    st.session_state.pdf_filename = data["nombre_pdf"] or ""
                    cols_order = ["Rol", "Cant", "Tiempo"] + [f"Precio {m}" for m in MARGINS] + [f"Subtotal {m}" for m in MARGINS]
                    records = json.loads(data["recursos_json"])
                    if records:
                        df_loaded = pd.DataFrame(records)
                        for c in [col for col in cols_order if col not in df_loaded.columns]:
                            df_loaded[c] = 0
                        st.session_state.items_df = df_loaded[cols_order]
                    else:
                        st.session_state.items_df = pd.DataFrame(columns=cols_order)
                    st.session_state.monederos_list = json.loads(data["monederos_json"])
                    pdf_path = data.get("pdf_path", "")
                    if pdf_path:
                        try:
                            pdf_bytes = _descargar_pdf_storage(pdf_path)
                            st.session_state.uploaded_pdf = _PDFProxy(pdf_bytes, data["nombre_pdf"] or pdf_path)
                            st.session_state.pdf_text = extraer_texto_pdf(pdf_bytes)
                        except Exception:
                            st.session_state.uploaded_pdf = None
                            st.session_state.pdf_text = ""
                    else:
                        st.session_state.uploaded_pdf = None
                        st.session_state.pdf_text = ""
                    st.session_state.folio_origen = data["folio"]
                    st.session_state.folio_actual = None
                    st.session_state.folio_preview = _generar_folio()
                    st.rerun()
                else:
                    st.error(f"No se encontró: **{folio_input.strip().upper()}**")
            else:
                st.warning("Ingresa un folio.")

    if st.session_state.folio_origen:
        st.info(
            f"Recotizando desde **{st.session_state.folio_origen}** — modifica y guarda para generar un folio nuevo.",
            icon="🔄"
        )
        if st.button("✖ Cancelar recotización", use_container_width=True, key="btn_limpiar_folio"):
            st.session_state.folio_origen = None
            st.session_state.folio_actual = None
            st.session_state.folio_preview = _generar_folio()
            st.rerun()

if not st.session_state.folio_origen:
    st.divider()

    # =========================
    # 1) Datos generales
    # =========================
    st.markdown("### 📄 Documentación y Enlaces")

    tipo_sel = st.radio(
        "👤 Tipo de cliente",
        options=["Externo", "Interno"],
        horizontal=True,
        index=0 if st.session_state.tipo_cliente == "Externo" else 1,
        key="radio_tipo_cliente"
    )
    if tipo_sel != st.session_state.tipo_cliente:
        st.session_state.tipo_cliente = tipo_sel
        st.rerun()

    col_ui1, col_ui2 = st.columns([1, 1], gap="large")

    with col_ui1:
        uploaded_file = st.file_uploader("📤 Subir PDF del Proyecto", type=["pdf"])
        if uploaded_file:
            if uploaded_file.name != st.session_state.pdf_filename:
                pdf_bytes = uploaded_file.getvalue()
                mb_orig = len(pdf_bytes) / (1024 * 1024)
                if len(pdf_bytes) > _LIMITE_PDF:
                    _aviso = st.info(
                        f"⏳ PDF de {mb_orig:.1f} MB detectado — comprimiendo, por favor espera…",
                        icon="🗜️",
                    )
                    with st.spinner("Comprimiendo PDF…"):
                        pdf_comprimido = _comprimir_pdf(pdf_bytes)
                    _aviso.empty()
                    mb_comp = len(pdf_comprimido) / (1024 * 1024)
                    st.session_state.uploaded_pdf = _PDFProxy(pdf_comprimido, uploaded_file.name)
                    if len(pdf_comprimido) <= _LIMITE_PDF:
                        st.success(f"✅ PDF comprimido: {mb_orig:.1f} MB → {mb_comp:.1f} MB")
                    else:
                        st.warning(
                            f"⚠️ El PDF se redujo a {mb_comp:.1f} MB pero aún supera los 50 MB. "
                            "Intenta exportar el PDF con menor resolución de imágenes."
                        )
                else:
                    st.session_state.uploaded_pdf = uploaded_file
                    st.success("✅ Archivo cargado correctamente")
                st.session_state.pdf_text = extraer_texto_pdf(st.session_state.uploaded_pdf.getvalue())
                st.session_state.pdf_filename = uploaded_file.name
            else:
                st.success("✅ Archivo cargado correctamente")

    with col_ui2:
        if st.session_state.tipo_cliente == "Externo":
            st.session_state.hubspot_link = st.text_input(
                "🔗 Enlace de HubSpot",
                value=st.session_state.hubspot_link,
                placeholder="https://app.hubspot.com/..."
            )
        else:
            st.session_state.hubspot_link = ""
            st.markdown("""
            <div class="badge-interno-box">
                <span style="color:#065F46;font-weight:600;">🏢 Cliente Interno</span>
                <p style="color:#047857;font-size:0.85rem;margin:0.3rem 0 0;">
                    No se requiere enlace de HubSpot.
                </p>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Validación de requisitos mínimos para continuar
    tiene_doc = st.session_state.uploaded_pdf is not None
    tiene_hubspot = st.session_state.tipo_cliente == "Interno" or st.session_state.hubspot_link.strip() != ""
    doc_completa = tiene_doc and tiene_hubspot

    if not doc_completa:
        if st.session_state.tipo_cliente == "Externo":
            st.info("📢 **Configuración Requerida:** Carga el PDF del proyecto y pega el enlace de HubSpot para continuar.", icon="🔒")
        else:
            st.info("📢 **Configuración Requerida:** Carga el PDF del proyecto para continuar.", icon="🔒")
        st.stop()

# =========================
# 2) Agregar recursos
# =========================
st.markdown("### 👥 Asignación de Recursos")

col_sel1, col_sel2 = st.columns(2)
with col_sel1:
    _mod_opts = ["DEDICADO", "STAFFING"]
    modalidad_sel = st.radio("🏷️ Modalidad", options=_mod_opts, horizontal=True,
                             index=_mod_opts.index(st.session_state.modalidad_global))
with col_sel2:
    _tar_opts = ["Mensual", "Por Hora"]
    tarifa_sel = st.radio("⏱️ Tipo de Tarifa", options=_tar_opts, horizontal=True,
                          index=_tar_opts.index(st.session_state.tarifa_global))

if modalidad_sel != st.session_state.modalidad_global or tarifa_sel != st.session_state.tarifa_global:
    if not st.session_state.items_df.empty:
        st.session_state.items_df = st.session_state.items_df.iloc[0:0]
        st.warning("⚠️ Configuración cambiada — la tabla de recursos fue reiniciada.", icon="🗑️")
    st.session_state.modalidad_global = modalidad_sel
    st.session_state.tarifa_global = tarifa_sel

colA, colB, colC = st.columns([1.5, 1, 1], gap="medium")

with colA:
    rol_sel = st.selectbox("👤 Perfil del Especialista", options=list(CATALOGO[st.session_state.modalidad_global].keys()))

precios = CATALOGO[st.session_state.modalidad_global][rol_sel]
if st.session_state.tarifa_global == "Por Hora":
    precios = [p / 160.0 for p in precios]

with colB:
    cantidad = st.number_input("Cantidad de personas", min_value=1, value=1)
    st.caption(f"Mínimo (21%): **${precios[0]:,.2f}**")

with colC:
    label_tiempo = "Meses" if st.session_state.tarifa_global == "Mensual" else "Horas"
    step_val = 0.5 if st.session_state.tarifa_global == "Mensual" else 1.0
    val_default = 1.0 if st.session_state.tarifa_global == "Mensual" else 160.0
    tiempo_val = st.number_input(label_tiempo, min_value=0.1, value=val_default, step=step_val)
    st.caption(f"Máximo (60%): **${precios[-1]:,.2f}**")

col_btn_add, col_btn_clear = st.columns([2, 1])
with col_btn_add:
    if st.button("➕ Agregar recurso", type="primary", use_container_width=True):
        factor = cantidad * tiempo_val
        data_nueva = {"Rol": rol_sel, "Cant": int(cantidad), "Tiempo": float(tiempo_val)}
        for i, m in enumerate(MARGINS):
            data_nueva[f"Precio {m}"] = precios[i]
            data_nueva[f"Subtotal {m}"] = math.ceil(precios[i] * factor)
        nuevo = pd.DataFrame([data_nueva])
        st.session_state.items_df = pd.concat([st.session_state.items_df, nuevo], ignore_index=True)
        st.rerun()
with col_btn_clear:
    if st.button("🗑️ Limpiar recursos", use_container_width=True):
        st.session_state.items_df = st.session_state.items_df.iloc[0:0]
        st.rerun()

# Calcular costo total de monederos (se usará en el resumen)
total_monederos_fee = sum(m["Total c/Fee"] for m in st.session_state.monederos_list)


# =========================
# 3) Detalle de Recursos
# =========================
st.markdown("### 📊 Detalle de Recursos")


# Tabla interactiva
label_tiempo_tabla = "Meses" if st.session_state.tarifa_global == "Mensual" else "Horas"
st.markdown(f"<p style='color: var(--text-muted); font-size: 0.95rem;'><em>Puedes editar directamente las Cantidades y {label_tiempo_tabla} en la siguiente tabla.</em></p>", unsafe_allow_html=True)

# Configurar visibilidad de columnas
column_config = {
    "Rol": st.column_config.TextColumn("Rol/Perfil", width="medium"),
    "Cant": st.column_config.NumberColumn("Cant.", min_value=1, step=1, width="small"),
    "Tiempo": st.column_config.NumberColumn(label_tiempo_tabla, min_value=0.1, step=0.5, width="small"),
}

# Ocultar columnas de Precio y configurar Subtotales visibles (21%, 25%, 60%)
for m in MARGINS:
    column_config[f"Precio {m}"] = None  # Ocultar siempre
    if m in ["21%", "25%", "60%"]:
        column_config[f"Subtotal {m}"] = st.column_config.NumberColumn(f"Subtotal {m}", format="$%.2f")
    else:
        column_config[f"Subtotal {m}"] = None  # Ocultar en la tabla UI

edited_df = st.data_editor(
    st.session_state.items_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config=column_config,
    key="editor_tabla"
)

if not edited_df.equals(st.session_state.items_df):
    st.session_state.items_df = recalcular(edited_df)
    st.rerun()

st.divider()

st.markdown("### 👛 Monederos")

incluir_monederos = st.toggle("Incluir Monederos en la cotización", value=False, key="toggle_monederos")

if incluir_monederos:
    colM1, colM2, colM3 = st.columns([1.5, 1, 1], gap="medium")

    with colM1:
        tipo_monedero = st.selectbox("🏦 Tipo de Monedero", options=list(MONEDEROS.keys()), key="sel_tipo_monedero")
        montos_disponibles = MONEDEROS[tipo_monedero]["Monto"]
        montos_con_fee = MONEDEROS[tipo_monedero]["Monto con fee"]
        fee_pct = "5%" if tipo_monedero == "Tiendas Neto" else "15%"

    with colM2:
        monto_idx = st.selectbox(
            "💵 Monto por Persona",
            options=range(len(montos_disponibles)),
            format_func=lambda i: f"${montos_disponibles[i]:,.0f}",
            key="sel_monto_monedero"
        )

    with colM3:
        personas_monedero = st.number_input("👤 Número de Personas", min_value=1, value=1, key="num_personas_monedero")
        costo_total_monedero = montos_con_fee[monto_idx] * personas_monedero
        st.success(f"Costo total **${costo_total_monedero:,.2f}**", icon="🧾")

    colBtnM, _ = st.columns([1, 2])
    with colBtnM:
        if st.button("➕ Agregar monedero al presupuesto", type="primary", use_container_width=True, key="btn_add_monedero"):
            st.session_state.monederos_list.append({
                "Tipo": tipo_monedero,
                "Monto Base": montos_disponibles[monto_idx],
                "Fee": fee_pct,
                "Monto c/Fee": round(montos_con_fee[monto_idx], 2),
                "Personas": int(personas_monedero),
                "Total c/Fee": round(costo_total_monedero, 2)
            })
            st.rerun()

    # Mostrar tabla de monederos agregados
    if st.session_state.monederos_list:
        st.markdown("<p style='color: var(--text-muted); font-size:0.9rem; margin-top:1rem;'><em>Monederos agregados a la cotización:</em></p>", unsafe_allow_html=True)
        df_monederos = pd.DataFrame(st.session_state.monederos_list)
        st.dataframe(df_monederos, use_container_width=True, hide_index=True)

        colLimpiaM, _ = st.columns([1, 4])
        with colLimpiaM:
            if st.button("🗑️ Limpiar monederos", use_container_width=True, key="btn_limpiar_monederos"):
                st.session_state.monederos_list = []
                st.rerun()
    else:
        st.info("No hay monederos agregados. Selecciona el tipo, monto y número de personas y presiona el botón.", icon="👛")
else:
    # Si el toggle está apagado, limpiar la lista para que no afecte los totales
    if st.session_state.monederos_list:
        st.session_state.monederos_list = []

st.divider()

# =========================
# 5) Resumen de Totales
# =========================
st.markdown("### 💹 Resumen de Totales")

# Cálculos finales
total_monederos_fee = sum(m["Total c/Fee"] for m in st.session_state.monederos_list)
totales = st.session_state.items_df[[f"Subtotal {m}" for m in MARGINS]].sum()

# Generar tarjetas dinámicamente solo para los márgenes seleccionados (21%, 25%, 60%)
cards_html = ""
for m in MARGINS:
    if m in ["21%", "25%", "60%"]:
        val_con_mon = totales[f"Subtotal {m}"] + total_monederos_fee
        m_num = m.replace("%", "")
        monedero_html = f'<div class="metric-detail">Monederos: ${total_monederos_fee:,.2f}</div>' if total_monederos_fee > 0 else ""
       
        cards_html += f"""
<div class="metric-container">
    <div class="metric-title">MARGEN {m}</div>
    <div class="metric-value val-{m_num}">${val_con_mon:,.2f}</div>
    <div class="metric-detail">Recursos: ${totales[f'Subtotal {m}']:,.2f}</div>
    {monedero_html}
</div>
"""

html_layout = f"""
<div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 1.2rem; margin-bottom: 2rem;">
    {cards_html}
</div>
"""
st.markdown(html_layout, unsafe_allow_html=True)

t_min = totales[f"Subtotal {MARGINS[0]}"] + total_monederos_fee
t_max = totales[f"Subtotal {MARGINS[-1]}"] + total_monederos_fee
st.warning(f"**⚠️ Regla de Negocio:** El total final (recursos + monederos) no debe ser menor (\${t_min:,.2f}) ({MARGINS[0]}) ni mayor (\${t_max:,.2f}) ({MARGINS[-1]})", icon="🚨")

st.divider()

# =========================
# 6) Exportar a Excel
# =========================
# generar_excel está definida al inicio del archivo

def enviar_correo(destinatario, asunto, cuerpo, adjuntos):
    remitente = st.secrets["email"]["cotizacion"]
    password = st.secrets["email"]["cotizacion_pass"]
   
   
    msg = MIMEMultipart()
    msg['From'] = remitente
    msg['To'] = destinatario
    msg['Subject'] = asunto
    msg.attach(MIMEText(cuerpo, 'plain'))

    for archivo_bytes, nombre_archivo in adjuntos:
        if archivo_bytes:
            # Determinar tipo MIME básico
            if nombre_archivo.lower().endswith('.xlsx'):
                main_type, sub_type = 'application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif nombre_archivo.lower().endswith('.pdf'):
                main_type, sub_type = 'application', 'pdf'
            else:
                main_type, sub_type = 'application', 'octet-stream'

            part = MIMEBase(main_type, sub_type)
            part.set_payload(archivo_bytes)
            encoders.encode_base64(part)
            # El método add_header maneja correctamente las comillas y evita espacios extras
            part.add_header('Content-Disposition', 'attachment', filename=nombre_archivo)
            msg.attach(part)

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(remitente, password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception:
        return False

def procesar_descarga_silenciosa(xlsx_data, file_name):
    lista_correos = [st.secrets["email"]["correo_1"], st.secrets["email"]["correo_2"]]
    hubspot_link = st.session_state.hubspot_link if st.session_state.hubspot_link else "No proporcionado"
    asunto = "Nueva Cotización Generada"
    cuerpo = f"Hola,\n\nSe ha generado una nueva cotización.\n\nLink de HubSpot: {hubspot_link}\n\nSaludos."

    adjuntos = [(xlsx_data, file_name)]
    if st.session_state.uploaded_pdf:
        adjuntos.append((st.session_state.uploaded_pdf.getvalue(), st.session_state.uploaded_pdf.name))

    for destinatario in lista_correos:
        enviar_correo(destinatario, asunto, cuerpo, adjuntos)

def _guardar_y_descargar(xlsx_data, file_name):
    nuevo_folio = st.session_state.folio_preview or _generar_folio()
    _guardar_en_db(nuevo_folio, folio_origen=st.session_state.folio_origen)
    st.session_state.folio_actual = nuevo_folio
    st.session_state.folio_origen = None
    procesar_descarga_silenciosa(xlsx_data, file_name)

st.markdown("### 📥 Generar Documentación")
st.markdown("<p style='color: var(--text-muted); font-size: 0.95rem;'>Agrega recursos para habilitar la descarga en Excel.</p>", unsafe_allow_html=True)

if not st.session_state.items_df.empty:
    xlsx_data = generar_excel(st.session_state.datos, st.session_state.items_df, st.session_state.monederos_list)
    fecha_str = date.today().strftime("%Y-%m-%d")
    folio_suffix = f"_{st.session_state.folio_actual}" if st.session_state.folio_actual else ""
    file_name = f"Cotizacion{folio_suffix}_{fecha_str}.xlsx"

    st.download_button(
        label="💾 Guardar cotización y Descargar Excel",
        data=xlsx_data,
        file_name=file_name,
        use_container_width=True,
        type="primary",
        on_click=_guardar_y_descargar,
        args=(xlsx_data, file_name)
    )
else:
    st.info("Para habilitar la descarga y el guardado, asegúrate de agregar al menos un recurso en la tabla.", icon="💡")

st.divider()
